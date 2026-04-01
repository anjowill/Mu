"""
Excel Exporter — produces a single .xlsx workbook with exactly 6 sheets.

Formatting:
- Row 1: Sheet title (merged, bold, 14pt, navy header)
- Row 2: Batch ID + generation timestamp (italic, grey)
- Row 3: Column headers (bold, navy background, white text)
- Data rows: alternating white / light-grey
- Freeze panes at row 4 (below headers)
- Text wrap on all cells
- Sheet 3: conditional coloring on total_tier column
- Sheets 4 & 5 (multi-table): blank row + bold sub-header between tables
"""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from schemas.sheet_schemas import (
    SheetOutput,
    Sheet4Output,
    Sheet5Output,
)

# ── Color constants ────────────────────────────────────────────────────────────
NAVY = "1F3864"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
AMBER = "FFC000"
GREEN = "70AD47"
RED = "FF0000"
TIER_COLORS = {
    "High-quality scale capital": GREEN,
    "Solid growth capital":       "92D050",
    "Mixed-quality execution capital": AMBER,
    "Low-conviction capital":     "FF7F00",
    "Weak / survival capital":    RED,
}

# ── Sheet column configs ───────────────────────────────────────────────────────
SHEET_CONFIGS = {
    "sheet_1": {
        "tab_name": "Use of Funds",
        "columns": [
            "Company Name",
            "What the company does (1 line)",
            "Reported use of funds (with citation)",
            "Use-of-funds classification",
            "Deal Size ($ Mn)",
            "Stage of Funding",
            "Industry",
            "Source Tier Used",
        ],
        "col_keys": [
            "company_name", "what_company_does", "reported_use_of_funds",
            "use_of_funds_classification", "deal_size_mn", "stage", "industry", "source_tier_used",
        ],
        "widths": [25, 40, 60, 35, 15, 20, 20, 18],
    },
    "sheet_2": {
        "tab_name": "Founder-Lens",
        "columns": [
            "Company",
            "Strategy Signal (Evidence + Citation)",
            "Capital Signal (Evidence + Citation)",
            "Execution Signal (Evidence + Citation)",
            "Founder Learning Insight",
        ],
        "col_keys": [
            "company", "strategy_signal", "capital_signal",
            "execution_signal", "founder_learning_insight",
        ],
        "widths": [22, 55, 55, 55, 55],
    },
    "sheet_3": {
        "tab_name": "Capital Quality",
        "columns": [
            "Name", "Stage", "Maturity", "Market", "Investors",
            "Stage Fit", "Purpose", "Total Score", "Total Tier", "Short Rationale",
        ],
        "col_keys": [
            "name", "stage", "maturity_score", "market_score", "investor_score",
            "stage_fit_score", "purpose_score", "total_score", "total_tier", "short_rationale",
        ],
        "widths": [25, 18, 10, 10, 12, 10, 10, 13, 30, 60],
    },
    "sheet_4": {
        "tab_name": "Sector Capital Map",
        "multi_table": True,
        "sub_tables": [
            {
                "title": "Table 1: Capital & Deal Count by Sector",
                "columns": ["Sector", "Deals", "% by Deal Count", "Capital ($ Mn)", "% of Capital"],
                "data_key": "sector_capital_table",
                "col_keys": ["sector", "deals", "pct_by_deal_count", "capital_mn", "pct_of_capital"],
                "widths": [30, 10, 18, 18, 15],
            },
            {
                "title": "Table 2: Sector Priority Ranking (By Capital Deployed)",
                "columns": ["Rank", "Sector", "Capital ($ Mn)", "% of Total", "Deals", "Priority Tier", "Interpretation"],
                "data_key": "sector_priority_table",
                "col_keys": ["rank", "sector", "capital_mn", "pct_of_total", "deals", "priority_tier", "interpretation"],
                "widths": [8, 30, 18, 13, 10, 14, 40],
            },
            {
                "title": "Table 3: Sector Scorecard",
                "columns": ["Sector", "Stage Profile", "Capital Quality", "Capital Intent", "Overall Character"],
                "data_key": "sector_scorecard_table",
                "col_keys": ["sector", "stage_profile", "capital_quality", "capital_intent", "overall_character"],
                "widths": [28, 25, 18, 25, 35],
            },
        ],
    },
    "sheet_5": {
        "tab_name": "Market Structure",
        "multi_table": True,
        "sub_tables": [
            {
                "title": "Table 2: Capital by Stage",
                "columns": ["Stage", "Capital ($ Mn)", "% of Total", "Deals"],
                "data_key": "capital_by_stage",
                "col_keys": ["stage", "capital_mn", "pct_of_total", "deals"],
                "widths": [30, 18, 14, 10],
            },
            {
                "title": "Table 3: B2C vs B2B Capital Split",
                "columns": ["Business Model", "Capital ($ Mn)", "% of Total", "Deals"],
                "data_key": "b2b_b2c_split",
                "col_keys": ["business_model", "capital_mn", "pct_of_total", "deals"],
                "widths": [22, 18, 14, 10],
            },
            {
                "title": "Table 4: Ticket Size Buckets",
                "columns": ["Ticket Size", "Deals", "Capital Share (%)"],
                "data_key": "ticket_size_buckets",
                "col_keys": ["ticket_size", "deals", "capital_share_pct"],
                "widths": [20, 10, 18],
            },
        ],
    },
    "sheet_6": {
        "tab_name": "Investor Intelligence",
        "columns": [
            "Investor",
            "Category",
            "Public Industry Focus / Sectors",
            "Portfolio Positioning",
            "Investment Thesis Points (cited)",
        ],
        "col_keys": [
            "investor", "category", "industry_focus", "portfolio_positioning", "thesis_points",
        ],
        "widths": [28, 16, 30, 40, 70],
    },
}


class ExcelExporter:
    def export(
        self,
        sheet_outputs: dict[str, SheetOutput],
        output_path: Path,
        batch_id: str,
    ) -> Path:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default empty sheet

        for sheet_id, cfg in SHEET_CONFIGS.items():
            output = sheet_outputs.get(sheet_id)
            if output is None:
                continue
            ws = wb.create_sheet(title=cfg["tab_name"])

            if cfg.get("multi_table"):
                self._write_multi_table_sheet(ws, output, cfg, batch_id)
            else:
                self._write_single_table_sheet(ws, output, cfg, batch_id)

        wb.save(output_path)
        return output_path

    # ── Single-table sheet (sheets 1, 2, 3, 6) ────────────────────────────────

    def _write_single_table_sheet(self, ws, output: SheetOutput, cfg: dict, batch_id: str) -> None:
        columns = cfg["columns"]
        col_keys = cfg["col_keys"]
        widths = cfg["widths"]
        sheet_id = output.sheet_id

        # Row 1: title
        _write_title_row(ws, cfg["tab_name"], len(columns), row=1)
        # Row 2: metadata
        _write_meta_row(ws, batch_id, output.generated_at, len(columns), row=2)
        # Row 3: headers
        _write_header_row(ws, columns, row=3)

        # Data rows
        for i, row_data in enumerate(output.rows):
            excel_row = 4 + i
            fill = PatternFill("solid", fgColor=LIGHT_GREY if i % 2 else WHITE)
            for col_idx, key in enumerate(col_keys, 1):
                val = row_data.get(key, "")
                cell = ws.cell(row=excel_row, column=col_idx, value=_safe_str(val))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = fill

                # Sheet 3: color total_tier column
                if sheet_id == "sheet_3" and key == "total_tier" and val:
                    color = TIER_COLORS.get(str(val), WHITE)
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(bold=True)

        _set_column_widths(ws, widths)
        ws.freeze_panes = "A4"

    # ── Multi-table sheet (sheets 4, 5) ───────────────────────────────────────

    def _write_multi_table_sheet(self, ws, output: SheetOutput, cfg: dict, batch_id: str) -> None:
        # Get the nested data structure (stored as first row in SheetOutput)
        nested = output.rows[0] if output.rows else {}
        max_cols = max(len(st["columns"]) for st in cfg["sub_tables"])

        _write_title_row(ws, cfg["tab_name"], max_cols, row=1)
        _write_meta_row(ws, batch_id, output.generated_at, max_cols, row=2)

        current_row = 3
        for sub in cfg["sub_tables"]:
            # Sub-table title row
            title_cell = ws.cell(row=current_row, column=1, value=sub["title"])
            title_cell.font = Font(bold=True, size=11)
            title_cell.fill = PatternFill("solid", fgColor="DCE6F1")
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(sub["columns"])
            )
            current_row += 1

            # Header row
            _write_header_row(ws, sub["columns"], row=current_row)
            current_row += 1

            # Data rows
            rows = nested.get(sub["data_key"], [])
            for i, row_data in enumerate(rows):
                fill = PatternFill("solid", fgColor=LIGHT_GREY if i % 2 else WHITE)
                for col_idx, key in enumerate(sub["col_keys"], 1):
                    val = row_data.get(key, "")
                    cell = ws.cell(row=current_row, column=col_idx, value=_safe_str(val))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.fill = fill
                current_row += 1

            # Blank separator row
            current_row += 1

        _set_column_widths(ws, cfg["sub_tables"][0]["widths"])
        ws.freeze_panes = "A3"


# ── Shared formatting helpers ──────────────────────────────────────────────────

def _write_title_row(ws, title: str, col_count: int, row: int) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=14, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if col_count > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=col_count,
        )
    ws.row_dimensions[row].height = 22


def _write_meta_row(ws, batch_id: str, generated_at: datetime, col_count: int, row: int) -> None:
    ts = generated_at.strftime("%Y-%m-%d %H:%M UTC") if generated_at else ""
    cell = ws.cell(row=row, column=1, value=f"Batch: {batch_id}  |  Generated: {ts}")
    cell.font = Font(italic=True, size=9, color="808080")
    if col_count > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=col_count,
        )


def _write_header_row(ws, columns: list[str], row: int) -> None:
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def _set_column_widths(ws, widths: list[int]) -> None:
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)
